from openpyxl import load_workbook
from string import ascii_lowercase
import itertools
import numpy as np
import numpy.matlib
from scipy.integrate import ode


def read_matrix(fname):
    wb = load_workbook(filename=fname)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    n_row = num_rows(ws)
    n_col = num_cols(ws)
    if n_row != n_col:
        raise Exception('Number of rows and columns is not consistant')
    m = np.array(read_all_rows(ws, 2, n_row, 2, n_col))
    return m


def read_vector(fname):
    wb = load_workbook(filename=fname)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    num_row = num_rows(ws, st_pt=1)
    output = np.array(read_all_rows(ws, 1, num_row, 2, 2))
    return output


def read_row(ws, row, st_pt, ed_pt):
    counter = 1
    output = []
    for s in iter_all_strings():
        if counter >= st_pt:
            output.append(ws[s + str(row)].value)
        if counter == ed_pt:
            return output
        counter += 1


def read_all_rows(ws, rows, rowe, st_pt, ed_pt):
    arr = []
    for row in range(rows, rowe + 1):
        c_row = read_row(ws, row, st_pt, ed_pt)
        c_row = none_2_zero(c_row)
        arr.append(c_row)
    return arr


def none_2_zero(arr):
    for i in range(len(arr)):
        if arr[i] is None:
            arr[i] = 0
    return arr


def num_rows(ws, st_pt=2):
    flag = True
    while flag:
        if ws['A' + str(st_pt)].value is None:
            return st_pt - 1
        else:
            st_pt += 1


def num_cols(ws, st_pt=2):
    count = 1
    for s in iter_all_strings():
        if count < st_pt:
            count += 1
            continue
        if ws[s + str(1)].value is None:
            return count - 1
        else:
            count += 1


def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_lowercase, repeat=size):
            yield "".join((s.upper() for s in s))
        size += 1


def equilibrium_state(amatrix, r_input):
    ia = np.linalg.inv(amatrix)
    n_eq = -np.matmul(ia, r_input)
    return n_eq


def calc_jacobian(A, r, n):
    i_matrix = np.eye(len(n))
    if len(n.shape) == 1:
        n_array = np.matlib.repmat(n, len(n), 1).T
    else:
        n_array = np.matlib.repmat(n, 1, len(n))
    J = i_matrix * r + A * n_array + i_matrix * np.matmul(A, n)

    return J


def calc_stability(A, r, n):
    J = calc_jacobian(A, r, n)
    ev = np.real(np.linalg.eig(J)[0])
    max_eig = np.max(ev)
    if max_eig < 0:
        return True
    else:
        return False


def draw_parameters(A, r):
    int_str = np.random.uniform(0, 1, np.shape(A))
    unc_int = np.abs(A) == 2
    unc_draws = np.random.uniform(0, 1, np.shape(A))
    A[(unc_draws < .5) & unc_int] = 0
    A[A == 2] = 1
    A[A == -2] = -1
    A_vals = (A * int_str)
    # A_vals = multiply_diag(A_vals)
    A_vals -= np.eye(np.shape(A_vals)[0])
    r_vals = np.random.uniform(0, 1, np.shape(r))
    return A_vals, r_vals


def gen_stable_param_set(A, r):
    flag = 0
    count = 0
    while flag == 0:
        At, rt = draw_parameters(A, r)
        n = equilibrium_state(At, rt)
        if np.all(n > 0):
            st = calc_stability(At, rt, n)
            if st:
                return At, rt, n
        count += 1
        print(count)


def multiply_diag(a, factor):
    a[np.eye(a.shape[0]) == 1] = a[np.eye(a.shape[0]) == 1] * factor
    return a


def remove_row_col(a_matrix, ind):
    a_shp = a_matrix.shape[0]
    if ind == 0:
        return a_matrix[1:a_shp, 1:a_shp]
    if ind == a_shp - 1:
        return A[0:a_shp - 1, 1:a_shp - 1]
    # If it is a 'central' option
    a_top_left = a_matrix[0:ind, 0:ind]
    a_top_right = a_matrix[0:ind, ind + 1:a_shp]
    a_bottom_left = a_matrix[ind + 1:a_shp, 0:ind]
    a_bottom_right = a_matrix[ind + 1:a_shp, ind + 1:a_shp]
    a_top = np.hstack([a_top_left, a_top_right])
    a_bottom = np.hstack([a_bottom_left, a_bottom_right])
    a_new = np.vstack([a_top, a_bottom])
    return a_new


def remove_rows_cols(A, r, inds):
    inds.sort()
    inds.reverse()
    An = A
    for ind in inds:
        An = remove_row_col(An, ind)
    rn = np.delete(r, inds)
    return An, rn


def add_rows_cols(A, r, n, inds, value=0):
    inds.sort()
    for ind in inds:
        A = np.insert(A, ind, value, axis=0)
        A = np.insert(A, ind, value, axis=1)
        r = np.insert(r, ind, value)
        n = np.insert(n, ind, value)
    return A, r, n


def gen_reduced_params(A, r, inds=None, reps=1):
    # inds is the species nums to remove
    A_output, r_output, n_output = [], [], []
    if inds is not None:
        A, r = remove_rows_cols(A, r, inds)
    for i in range(reps):
        Ap, rp, Np = gen_stable_param_set(A, r)
        if inds is not None:
            Ap, rp, Np = add_rows_cols(Ap, rp, Np, inds)
        A_output.append(Ap)
        r_output.append(rp)
        n_output.append(Np)
    A_output = np.array(A_output)
    r_output = np.array(r_output)
    n_output = np.array(n_output)
    return A_output, r_output, n_output


def de(t, y, A, r):
    return r * y + np.matmul(A, y) * y


def de_solve(T, y, A, r):
    rde = ode(de).set_integrator('zvode', method='bdf', with_jacobian=False)
    rde.set_initial_value(y, 0).set_f_params(A, r)
    return np.real(rde.integrate(T))


A = read_matrix('Phillip_islands_community.xlsx')
r = read_vector('Phillip_islands_r.xlsx')
n = equilibrium_state(A, r)
J = calc_jacobian(A, r, n)
st = calc_stability(A, r, n)
# print(st)
print(gen_stable_param_set(A, r))
# remove_rows_cols(A,r,[2,5])
# print(gen_reduced_params(A,r,[2,5],5))

Ap, rp, Np = gen_reduced_params(A, r, [2, 5], 1)
Ap = Ap[0]
rp = rp[0]
Np = Np[0]
Np[0] = 0
New_N = de_solve(1, Np, Ap, rp)
