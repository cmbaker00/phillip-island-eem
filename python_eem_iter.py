from openpyxl import load_workbook
from string import ascii_lowercase
import itertools
import numpy as np
import numpy.matlib
from scipy.integrate import ode
import copy
import pickle
import matplotlib.pyplot as plt
import copy

def read_matrix(fname):
    wb = load_workbook(filename=fname)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    n_row = num_rows(ws)
    n_col = num_cols(ws)
    if n_row != n_col:
        raise Exception('Number of rows and columns is not consistant')
    m = np.array(read_all_rows(ws, 2, n_row, 2, n_col))
    return m


def read_vector(fname):
    wb = load_workbook(filename=fname)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    num_row = num_rows(ws, st_pt=1)
    output = np.array(read_all_rows(ws, 1, num_row, 2, 2))
    return output


def read_row(ws, row, st_pt, ed_pt):
    counter = 1
    output = []
    for s in iter_all_strings():
        if counter >= st_pt:
            output.append(ws[s + str(row)].value)
        if counter == ed_pt:
            return output
        counter += 1


def read_all_rows(ws, rows, rowe, st_pt, ed_pt):
    arr = []
    for row in range(rows, rowe + 1):
        c_row = read_row(ws, row, st_pt, ed_pt)
        c_row = none_2_zero(c_row)
        arr.append(c_row)
    return arr


def none_2_zero(arr):
    for i in range(len(arr)):
        if arr[i] is None:
            arr[i] = 0
    return arr


def num_rows(ws, st_pt=2):
    flag = True
    while flag:
        if ws['A' + str(st_pt)].value is None:
            return st_pt - 1
        else:
            st_pt += 1


def num_cols(ws, st_pt=2):
    count = 1
    for s in iter_all_strings():
        if count < st_pt:
            count += 1
            continue
        if ws[s + str(1)].value is None:
            return count - 1
        else:
            count += 1


def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_lowercase, repeat=size):
            yield "".join((s.upper() for s in s))
        size += 1


def equilibrium_state(amatrix, r_input):
    ia = np.linalg.inv(amatrix)
    n_eq = -np.matmul(ia, r_input)
    return n_eq


def calc_jacobian(A, r, n):
    i_matrix = np.eye(len(n))
    if len(n.shape) == 1:
        n_array = np.matlib.repmat(n, len(n), 1).T
    else:
        n_array = np.matlib.repmat(n, 1, len(n))
    J = i_matrix * r + A * n_array + i_matrix * np.matmul(A, n)

    return J


def calc_stability(A, r, n):
    J = calc_jacobian(A, r, n)
    ev = np.real(np.linalg.eig(J)[0])
    max_eig = np.max(ev)
    if max_eig < 0:
        return True
    else:
        return False


def draw_parameters(A, r):
    int_str = np.random.uniform(0, 1, np.shape(A))
    unc_int = np.abs(A) == 2
    unc_draws = np.random.uniform(0, 1, np.shape(A))
    A[(unc_draws < .5) & unc_int] = 0
    A[A == 2] = 1
    A[A == -2] = -1
    A_vals = (A * int_str)
    # A_vals = multiply_diag(A_vals)
    A_vals -= np.eye(np.shape(A_vals)[0])
    r_vals = np.random.uniform(0, 1, np.shape(r))
    return A_vals, r_vals


def gen_stable_param_set(A, r):
    flag = 0
    count = 0
    while flag == 0:
        At, rt = draw_parameters(A, r)
        n = equilibrium_state(At, rt)
        if np.all(n > 0):
            st = calc_stability(At, rt, n)
            if st:
                return At, rt, n
        count += 1
        # print(count)


def multiply_diag(a, factor):
    a[np.eye(a.shape[0]) == 1] = a[np.eye(a.shape[0]) == 1] * factor
    return a


def remove_row_col(a_matrix, ind):
    a_shp = a_matrix.shape[0]
    if ind == 0:
        return a_matrix[1:a_shp, 1:a_shp]
    if ind == a_shp - 1:
        return a_matrix[0:a_shp - 1, 1:a_shp - 1]
    # If it is a 'central' option
    a_top_left = a_matrix[0:ind, 0:ind]
    a_top_right = a_matrix[0:ind, ind + 1:a_shp]
    a_bottom_left = a_matrix[ind + 1:a_shp, 0:ind]
    a_bottom_right = a_matrix[ind + 1:a_shp, ind + 1:a_shp]
    a_top = np.hstack([a_top_left, a_top_right])
    a_bottom = np.hstack([a_bottom_left, a_bottom_right])
    a_new = np.vstack([a_top, a_bottom])
    return a_new


def remove_rows_cols(A, r, inds):
    inds.sort()
    inds.reverse()
    An = A
    for ind in inds:
        An = remove_row_col(An, ind)
    rn = np.delete(r, inds)
    return An, rn


def add_rows_cols(A, r, n, inds, value=0):
    inds.sort()
    for ind in inds:
        A = np.insert(A, ind, value, axis=0)
        A = np.insert(A, ind, value, axis=1)
        r = np.insert(r, ind, value)
        n = np.insert(n, ind, value)
    return A, r, n

def gen_other_params(Ap, A_input, rp):
    A_draw, r_draw = draw_parameters(A_input, rp)
    locs = (A_input != 0) == (Ap == 0)
    rlocs = rp == 0
    Ap[locs] = A_draw[locs]
    rp[rlocs] = r_draw[rlocs]
    return Ap, rp

def gen_reduced_params(A_input, r, inds=None, reps=1):
    # inds is the species nums to remove
    A_output, r_output, n_output = [], [], []
    if inds is not None:
        A, r = remove_rows_cols(A_input, r, inds)
    else:
        A = A_input
    for i in range(reps):
        Ap, rp, Np = gen_stable_param_set(A, r)
        if inds is not None:
            Ap, rp, Np = add_rows_cols(Ap, rp, Np, inds)
            Ap, rp = gen_other_params(Ap, A_input, rp)
        A_output.append(Ap)
        r_output.append(rp)
        n_output.append(Np)
    A_output = np.array(A_output)
    r_output = np.array(r_output)
    n_output = np.array(n_output)
    return A_output, r_output, n_output


def de(t, y, A, r):
    return r * y + np.matmul(A, y) * y


def de_solve(T, y, A, r):
    rde = ode(de).set_integrator('lsoda', method='bdf', with_jacobian=False)
    rde.set_initial_value(y, 0).set_f_params(A, r)
    return np.real(rde.integrate(T))

def de_fix(t, y, A, r, f_id):
    rates = r * y + np.matmul(A, y) * y
    rates[f_id] = 0
    return rates


def de_solve_fix(T, y, A, r, f_id):
    rde = ode(de_fix).set_integrator('lsoda', method='bdf', with_jacobian=False)
    rde.set_initial_value(y, 0).set_f_params(A, r, f_id)
    return np.real(rde.integrate(T))


class EEM:
    def __init__(self, a, r, rem=None, max_sets=np.inf):
        self.a = a
        self.r = r
        self.rem = rem
        self.max = max_sets
        self.counter = 0

    def __iter__(self):
        return self

    def __next__(self):
        if self.counter >= self.max:
            raise StopIteration
        Ap, rp, Np = gen_reduced_params(self.a, self.r, self.rem, 1)
        self.counter += 1
        return Ap[0], rp[0], Np[0]

class EEM_rem:
    def __init__(self, EEM_gen, removal, response, max_iter=np.inf):
        self.EEM_gen = EEM_gen
        self.rem = removal
        self.resp = np.array(response)
        self.count = 0
        self.max = max_iter

    def __iter__(self):
        iter(self.EEM_gen)
        return self

    def __next__(self):
        if self.count >= self.max:
            raise StopIteration
        flag = 0
        c = 0
        while flag == 0:
            c += 1
            # print(c)
            a, r, n = next(self.EEM_gen)
            n_change = copy.copy(n)
            n_change[self.rem] = 0
            n_new = de_solve(.1, n_change, a, r)
            change = n_new > n
            # print(change[self.resp[:,0]])
            if (change[self.resp[:, 0]] == self.resp[:, 1]).all():
                flag = 1
        self.count += 1
        # print(a, r, n)
        return a, r, n

class EEM_stable_from_prev_conds:
    def __init__(self, EEM_gen, removal, max_iter=np.inf):
        self.EEM_gen = EEM_gen
        self.rem = removal
        self.count = 0
        self.max = max_iter

    def __iter__(self):
        iter(self.EEM_gen)
        return self

    def __next__(self):
        if self.count >= self.max:
            raise StopIteration
        flag = 0
        while flag == 0:
            a, r, n = next(self.EEM_gen)
            a_rem, r_rem = remove_rows_cols(a, r, self.rem)
            n_rem = equilibrium_state(a_rem, r_rem)
            if calc_stability(a_rem, r_rem, n_rem):
                for ind in np.sort(self.rem): #reinsert the appropriate zeros
                    n_rem = np.insert(n_rem, ind, 0)
                self.count += 1
                return a, r, n_rem

class EEM_reintro:
    def __init__(self, ensemble, reintro_id, control_id, control_amount, T = 10):
        self.ensemble = ensemble
        self.reintro = reintro_id
        self.control = control_id
        self.control_level = control_amount
        self.max = len(ensemble)
        self.count = 0
        self.outcomes = []
        self.T = T

    def __iter__(self):
        return self


    def __next__(self):
        if self.count >= self.max:
            raise StopIteration
        ratio = self.get_outcome(self.count)
        self.outcomes.append(ratio)
        self.count += 1
        return ratio

    def get_outcome(self, pset):
        if pset < len(self.outcomes):
            return self.outcomes[pset]
        param_set = self.ensemble[pset]
        a = param_set[0]
        r = param_set[1]
        n = param_set[2]
        n_old = copy.copy(n)
        start_abund = np.min(n[n != 0])
        if len(self.reintro) > 0:
            n[self.reintro] = start_abund
        n_old = copy.copy(n)
        if len(self.control) > 0:
            n[self.control] = n[self.control]*self.control_level
        n_new = de_solve_fix(self.T, n, a, r, self.control)
        ratio = n_new/n_old
        ratio[np.isnan(ratio)] = 0
        return ratio




def save_object(obj, filename):
    with open(filename, 'wb') as output:  # Overwrites any existing file.
        pickle.dump(obj, output, pickle.HIGHEST_PROTOCOL)

def load_object(filename):
    with open(filename, 'rb') as input:
        return pickle.load(input)

def generate_phillip_island_ensemble(fname='PI_EEM', rep=10000, response = [], rem_node = [], final_removed=[]):
    a_input = read_matrix('Phillip_islands_community.xlsx').transpose()
    r_input = read_vector('Phillip_islands_r.xlsx')
    gen_init_stable = EEM(a_input, r_input, [2, 5])
    gen_cond_params = EEM_rem(gen_init_stable, rem_node, response)

    gen_final_param_set = EEM_stable_from_prev_conds(gen_cond_params, final_removed, max_iter=rep)
    param_sets = [[params[0], params[1], params[2]] for params in gen_final_param_set]
    fname = fname + '_' + str(rep) + '.pkl'
    save_object(param_sets, fname)

if __name__ == "__main__":
    reps = 1000
    rem_node = [0] # node to be removed
    response = [[7, True], [10, True], [15, True]] # these increase when node is removed
    final_removed = [0, 16] # return an ensemble without these species
    # generate_phillip_island_ensemble(rep=reps, response = [[7, True], [10, True], [15, True]], rem_node = [0],final_removed = [0, 16])
    ensemble = load_object('PI_EEM_{0}.pkl'.format(reps))
    control = [1]  # controlling cats
    control_level = .5 # control cats to 50% of current
    reintro_sp = [16]  # reintroduce bandicoots
    reintro = EEM_reintro(ensemble, reintro_sp, control, control_level)
    bandicoot_data = [abund[16] for abund in reintro]
    mean_increase = np.mean(np.array(bandicoot_data)>1)
    plt.hist(bandicoot_data, bins='auto')
    plt.title("Frequency of increase: {}".format((int(mean_increase*1000))/1000))
    plt.show()